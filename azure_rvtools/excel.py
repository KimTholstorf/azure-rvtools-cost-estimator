"""Azure Calculator-style Excel workbook output for azure-rvtools."""
from __future__ import annotations

from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from .output import ReservationRec, VMResult

VERSION = "1.0.2"

# ---------------------------------------------------------------------------
# Azure Support Plan pricing (monthly USD, fixed — not region-specific)
# Source: https://azure.microsoft.com/en-us/support/plans/
# ---------------------------------------------------------------------------
SUPPORT_PLANS: dict[str, tuple[str, float]] = {
    #  cli key          display name              monthly USD
    "basic":            ("Basic",                 0.0),
    "developer":        ("Developer",             29.0),
    "standard":         ("Standard",              100.0),
    "professional-direct": ("Professional Direct", 1_000.0),
}

# ---------------------------------------------------------------------------
# Colours (hex, no leading #)
# ---------------------------------------------------------------------------
_BLUE_HEADER   = "DDEBF7"   # Azure Calculator header row light blue
_GRAY_DISC     = "D3D3D3"   # Disclaimer section background
_GRAY_TOTAL    = "F2F2F2"   # Total row tint
_GRAY_ALT      = "F9F9F9"   # Alternating row tint for detail sheets
_BLUE_TITLE    = "1F4E79"   # Dark Azure blue for title text
_WHITE         = "FFFFFF"

# ---------------------------------------------------------------------------
# Reusable style factories  (openpyxl styles are not thread-safe to share)
# ---------------------------------------------------------------------------
def _font(size=11, bold=False, color="000000", italic=False):
    return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _align(horizontal="left", vertical="top", wrap=False):
    return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap)

# Currency number format matching the Azure Calculator export
_CURRENCY_FMT = '[$$]#,##0.00'
_INT_FMT      = '#,##0'

# ---------------------------------------------------------------------------
# Region display name mapping
# ---------------------------------------------------------------------------
_REGION_NAMES: dict[str, str] = {
    "westeurope": "West Europe", "northeurope": "North Europe",
    "eastus": "East US", "eastus2": "East US 2",
    "westus": "West US", "westus2": "West US 2", "westus3": "West US 3",
    "centralus": "Central US", "southcentralus": "South Central US",
    "northcentralus": "North Central US", "westcentralus": "West Central US",
    "canadacentral": "Canada Central", "canadaeast": "Canada East",
    "brazilsouth": "Brazil South",
    "uksouth": "UK South", "ukwest": "UK West",
    "francecentral": "France Central", "francesouth": "France South",
    "germanywestcentral": "Germany West Central", "germanynorth": "Germany North",
    "switzerlandnorth": "Switzerland North", "switzerlandwest": "Switzerland West",
    "norwayeast": "Norway East", "norwaywest": "Norway West",
    "swedencentral": "Sweden Central", "polandcentral": "Poland Central",
    "australiaeast": "Australia East", "australiasoutheast": "Australia Southeast",
    "australiacentral": "Australia Central",
    "eastasia": "East Asia", "southeastasia": "Southeast Asia",
    "japaneast": "Japan East", "japanwest": "Japan West",
    "koreacentral": "Korea Central", "koreasouth": "Korea South",
    "centralindia": "Central India", "westindia": "West India",
    "southindia": "South India",
    "southafricanorth": "South Africa North", "southafricawest": "South Africa West",
    "uaenorth": "UAE North", "uaecentral": "UAE Central",
}

def _region_display(region: str) -> str:
    return _REGION_NAMES.get(region.lower(), region)


# ---------------------------------------------------------------------------
# Helper: apply a style dict to a cell
# ---------------------------------------------------------------------------
def _style(cell, font=None, fill=None, align=None, number_format=None):
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if align:
        cell.alignment = align
    if number_format:
        cell.number_format = number_format


# ---------------------------------------------------------------------------
# Sheet 1 — Azure Calculator look-alike "Estimate"
# ---------------------------------------------------------------------------

def _vm_description(sku_name: str, vcpus: int, ram_gb: int, os_type: str,
                    count: int, pricing_mode: str, reserved_term: str = "3-year",
                    hybrid_benefit: bool = False,
                    is_reserved: bool = False) -> str:
    if hybrid_benefit:
        os_label = "Linux (Azure Hybrid Benefit)"
    elif os_type == "windows":
        os_label = "Windows (License included)"
    else:
        os_label = "Linux"
    if pricing_mode == "all-reserved" or (pricing_mode == "realistic" and is_reserved):
        years = "3 year" if reserved_term == "3-year" else "1 year"
        term = f"({years} reserved)"
    else:
        term = "(pay as you go)"
    return (
        f"{count:,} {sku_name} ({vcpus} vCPUs, {ram_gb} GB RAM) {term}, "
        f"{os_label}, 730 Hours"
    )


def _disk_description(results: list[VMResult], disk_type: str) -> str:
    tier_counts: dict[str, int] = defaultdict(int)
    total = 0
    for r in results:
        if r.vm.is_powered_on:
            for tier, cnt in r.disk_tiers:
                tier_counts[tier] += cnt
                total += cnt

    def _tier_num(t: str) -> int:
        import re
        m = re.search(r"\d+", t)
        return int(m.group()) if m else 0

    detail = ", ".join(
        f"{t}: {c:,}"
        for t, c in sorted(tier_counts.items(), key=lambda x: _tier_num(x[0]))
    )
    type_label = {"premium-ssd": "Premium SSD", "standard-ssd": "Standard SSD",
                  "standard-hdd": "Standard HDD"}.get(disk_type, disk_type)
    return f"{total:,} {type_label} managed disks — {detail}"


def _build_estimate_sheet(
    wb: Workbook,
    results: list[VMResult],
    recommendations: list[ReservationRec],
    region: str,
    currency: str,
    pricing_mode: str,
    disk_type: str,
    rvtools_filename: str,
    reserved_term: str = "3-year",
    hybrid_benefit: bool = False,
    support_plan: str = "basic",
    realistic_top_skus: frozenset[str] = frozenset(),
    disk_source: str = "provisioned",
) -> None:
    ws = wb.active
    ws.title = "Estimate"

    region_display = _region_display(region)
    now = datetime.now(timezone.utc)

    # --- Column widths (matching Azure Calculator) ---
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 62
    ws.column_dimensions["F"].width = 26

    # --- Row 1: Title ---
    ws.row_dimensions[1].height = 24
    ws["A1"] = "Microsoft Azure Estimate"
    _style(ws["A1"],
           font=_font(size=14, bold=True, color=_BLUE_TITLE),
           align=_align("left", "center"))
    ws.merge_cells("A1:C1")

    # --- Row 2: Subtitle ---
    ws.row_dimensions[2].height = 18
    ws["A2"] = "Your Estimate"
    _style(ws["A2"],
           font=_font(size=12, bold=True),
           align=_align("left", "center"))
    ws.merge_cells("A2:C2")

    # --- Row 3: blank ---
    ws.row_dimensions[3].height = 6

    # --- Row 4: Column headers ---
    ws.row_dimensions[4].height = 20
    headers = [
        ("A4", "Service category"),
        ("B4", "Service type"),
        ("C4", "Custom name"),
        ("D4", "Region"),
        ("E4", "Description"),
        ("F4", "Estimated monthly cost"),
    ]
    for coord, label in headers:
        cell = ws[coord]
        cell.value = label
        _style(cell,
               font=_font(size=11, bold=True),
               fill=_fill(_BLUE_HEADER),
               align=_align("left", "top", wrap=True))

    # --- Group VMs by (sku_name, os_type) for estimate rows ---
    # sku_groups: (sku_name, os_type) → list of VMResult
    sku_groups: dict[tuple[str, str], list[VMResult]] = defaultdict(list)
    for r in results:
        if r.vm.is_powered_on and r.sku is not None:
            sku_groups[(r.sku.name, r.vm.os_type)].append(r)

    # Sort by total PAYG cost descending (largest first, like calculator)
    def _group_sort_key(item: tuple) -> float:
        return sum(r.payg_compute_monthly for r in item[1])

    sorted_groups = sorted(sku_groups.items(), key=_group_sort_key, reverse=True)

    data_start_row = 5
    current_row = data_start_row

    vm_row_refs: list[str] = []   # collect F-cell addresses for SUM formula

    for (sku_name, os_type), group in sorted_groups:
        sku = group[0].sku
        count = len(group)

        is_reserved = (
            pricing_mode == "all-reserved"
            or (pricing_mode == "realistic" and sku_name in realistic_top_skus)
        )

        if is_reserved:
            monthly = sum(
                r.reserved_compute_monthly for r in group
                if r.reserved_compute_monthly is not None
            )
        else:
            monthly = sum(r.payg_compute_monthly for r in group)

        desc = _vm_description(
            sku_name, sku.vcpus, sku.ram_gb, os_type, count, pricing_mode,
            reserved_term=reserved_term, hybrid_benefit=hybrid_benefit,
            is_reserved=is_reserved,
        )

        ws.row_dimensions[current_row].height = 48
        ws.cell(current_row, 1, "Compute")
        ws.cell(current_row, 2, "Virtual Machines")
        ws.cell(current_row, 3, "")
        ws.cell(current_row, 4, region_display)
        ws.cell(current_row, 5, desc)
        f_cell = ws.cell(current_row, 6, monthly)
        f_cell.number_format = _CURRENCY_FMT

        for col in range(1, 7):
            _style(ws.cell(current_row, col),
                   font=_font(11),
                   align=_align("left", "top", wrap=(col in (4, 5))))
        f_cell.alignment = _align("left", "top")

        vm_row_refs.append(f"F{current_row}")
        current_row += 1

    # --- Disk row ---
    total_disk = sum(r.disk_monthly for r in results if r.vm.is_powered_on)
    disk_desc  = _disk_description(results, disk_type)

    ws.row_dimensions[current_row].height = 48
    ws.cell(current_row, 1, "Storage")
    ws.cell(current_row, 2, "Managed Disks")
    ws.cell(current_row, 3, "")
    ws.cell(current_row, 4, region_display)
    ws.cell(current_row, 5, disk_desc)
    disk_f = ws.cell(current_row, 6, total_disk)
    disk_f.number_format = _CURRENCY_FMT

    for col in range(1, 7):
        _style(ws.cell(current_row, col),
               font=_font(11),
               align=_align("left", "top", wrap=(col in (4, 5))))
    disk_f.alignment = _align("left", "top")

    vm_row_refs.append(f"F{current_row}")
    current_row += 1

    # --- Support row ---
    support_display, support_price = SUPPORT_PLANS.get(support_plan, ("Basic", 0.0))
    support_desc = f"Azure Support – {support_display} plan"
    ws.row_dimensions[current_row].height = 17
    ws.cell(current_row, 1, "Support")
    ws.cell(current_row, 4, region_display)
    ws.cell(current_row, 5, support_desc)
    sup_f = ws.cell(current_row, 6, support_price)
    sup_f.number_format = _CURRENCY_FMT
    for col in range(1, 7):
        _style(ws.cell(current_row, col), font=_font(11), align=_align("left", "top", wrap=(col == 5)))
    sup_f.alignment = _align("left", "top")
    vm_row_refs.append(f"F{current_row}")
    current_row += 1

    # --- Blank separator ---
    ws.row_dimensions[current_row].height = 8
    current_row += 1

    # --- Licensing Program ---
    ws.row_dimensions[current_row].height = 17
    ws.cell(current_row, 4, "Licensing Program")
    ws.cell(current_row, 5, "Microsoft Customer Agreement (MCA)")
    _style(ws.cell(current_row, 4), font=_font(11), align=_align("left", "top"))
    _style(ws.cell(current_row, 5), font=_font(11), align=_align("left", "top"))
    current_row += 1

    # --- Billing Account / Profile (blank, match format) ---
    for label in ("Billing Account", "Billing Profile"):
        ws.row_dimensions[current_row].height = 17
        ws.cell(current_row, 4, label)
        ws.cell(current_row, 5, "")
        _style(ws.cell(current_row, 4), font=_font(11), align=_align("left", "top"))
        current_row += 1

    # --- Total row ---
    ws.row_dimensions[current_row].height = 17
    ws.cell(current_row, 4, "Total")
    total_cell = ws.cell(current_row, 6)
    total_cell.value = f"=SUM({','.join(vm_row_refs)})"
    total_cell.number_format = _CURRENCY_FMT

    for col in range(1, 7):
        c = ws.cell(current_row, col)
        _style(c, font=_font(11, bold=True), fill=_fill(_GRAY_TOTAL),
               align=_align("left", "top"))
    total_row_num = current_row
    current_row += 1

    # --- Blank ---
    ws.row_dimensions[current_row].height = 8
    current_row += 1

    # --- Disclaimer header ---
    ws.row_dimensions[current_row].height = 17
    ws.cell(current_row, 1, "Disclaimer")
    _style(ws.cell(current_row, 1), font=_font(11, bold=True), align=_align())
    current_row += 1

    # --- Disclaimer body (gray, merged) ---
    disc_text = (
        "All prices shown are in United States – Dollar ($) USD. "
        "This is a summary estimate, not a quote. "
        "For up to date pricing information please visit https://azure.microsoft.com/pricing/calculator/"
    )
    ws.row_dimensions[current_row].height = 36
    ws.cell(current_row, 1, disc_text)
    _style(ws.cell(current_row, 1),
           font=_font(10), fill=_fill(_GRAY_DISC),
           align=_align("left", "center", wrap=True))
    ws.merge_cells(f"A{current_row}:F{current_row}")
    current_row += 1

    # --- Generation line ---
    if pricing_mode == "all-reserved":
        pricing_label = ("3-Year Reserved" if reserved_term == "3-year" else "1-Year Reserved")
    elif pricing_mode == "realistic":
        rsv_label = "3-Year Reserved" if reserved_term == "3-year" else "1-Year Reserved"
        pricing_label = f"Realistic (top {len(realistic_top_skus)} SKUs at {rsv_label}, rest PAYG)"
    else:
        pricing_label = "Pay as you go"
    ahb_note = " Azure Hybrid Benefit applied." if hybrid_benefit else ""
    disk_source_label = "in-use capacity" if disk_source == "in-use" else "provisioned capacity"
    gen_text = (
        f"This estimate was generated by azure-rvtools v{VERSION} "
        f"from '{rvtools_filename}' on "
        f"{now.strftime('%Y-%m-%d %H:%M')} UTC. "
        f"Region: {region_display}. "
        f"Pricing: {pricing_label}.{ahb_note} "
        f"Disk source: {disk_source_label}."
    )
    ws.row_dimensions[current_row].height = 24
    ws.cell(current_row, 1, gen_text)
    _style(ws.cell(current_row, 1),
           font=_font(10, italic=True), fill=_fill(_GRAY_DISC),
           align=_align("left", "center", wrap=True))
    ws.merge_cells(f"A{current_row}:F{current_row}")

    # Freeze panes below header row
    ws.freeze_panes = "A5"


# ---------------------------------------------------------------------------
# Sheet 2 — VM Detail
# ---------------------------------------------------------------------------

def _build_vm_detail_sheet(
    wb: Workbook,
    results: list[VMResult],
    pricing_mode: str,
    region: str,
    reserved_term: str = "3-year",
) -> None:
    from .output import format_disk_summary, _rsv_label

    ws = wb.create_sheet("VM Detail")

    col_widths = [36, 8, 10, 24, 10, 34, 18, 18, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    show_reserved = pricing_mode in ("all-reserved", "realistic")

    headers = ["VM Name", "vCPUs", "RAM (GB)", "Azure SKU", "OS",
               "Disks", "PAYG/mo (USD)"]
    if show_reserved:
        headers.append(f"{_rsv_label(reserved_term)} Rsv/mo (USD)")
    headers.append("Notes")

    # Header row
    ws.row_dimensions[1].height = 20
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        _style(cell,
               font=_font(11, bold=True),
               fill=_fill(_BLUE_HEADER),
               align=_align("left", "center"))

    ws.freeze_panes = "A2"

    for row_idx, r in enumerate(results, 2):
        ws.row_dimensions[row_idx].height = 15
        fill = _fill(_GRAY_ALT) if row_idx % 2 == 0 else None

        notes_parts = []
        if not r.vm.is_powered_on:
            notes_parts.append("Powered off")
        if r.sku is None:
            notes_parts.append("No matching SKU")
        else:
            notes_parts.extend(r.sku_notes)

        ram = r.vm.ram_gb
        ram_display = int(ram) if ram == int(ram) else round(ram, 1)

        row_data = [
            r.vm.name,
            r.vm.vcpus,
            ram_display,
            r.sku.name if r.sku else "—",
            r.vm.os_type,
            format_disk_summary(r.disk_tiers),
            r.payg_total_monthly if r.sku else None,
        ]
        if show_reserved:
            row_data.append(r.reserved_total_monthly)
        row_data.append("; ".join(notes_parts))

        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row_idx, col_idx, val)
            is_money = (col_idx == 7) or (show_reserved and col_idx == 8)
            if is_money and isinstance(val, (int, float)):
                cell.number_format = _CURRENCY_FMT
                _style(cell, font=_font(10), fill=fill,
                       align=_align("right", "center"))
            else:
                _style(cell, font=_font(10), fill=fill,
                       align=_align("left", "center"))


# ---------------------------------------------------------------------------
# Sheet 3 — Reservation Recommendations
# ---------------------------------------------------------------------------

def _build_reservations_sheet(
    wb: Workbook,
    recommendations: list[ReservationRec],
    total_disk_monthly: float,
    reserved_term: str = "3-year",
) -> None:
    from .output import _rsv_label_long

    ws = wb.create_sheet("Reservations")

    col_widths = [28, 8, 22, 22, 18, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    rsv_col_header = f"{_rsv_label_long(reserved_term)}/mo (USD)"
    headers = ["SKU", "VMs", "PAYG Compute/mo (USD)",
               rsv_col_header, "Saves/mo (USD)", "Saves %"]

    ws.row_dimensions[1].height = 20
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        _style(cell,
               font=_font(11, bold=True),
               fill=_fill(_BLUE_HEADER),
               align=_align("left", "center"))

    ws.freeze_panes = "A2"

    for row_idx, rec in enumerate(recommendations, 2):
        ws.row_dimensions[row_idx].height = 15
        fill = _fill(_GRAY_ALT) if row_idx % 2 == 0 else None

        row_data = [
            rec.sku_name, rec.vm_count,
            rec.payg_compute_monthly, rec.rsv_compute_monthly,
            rec.savings_monthly, rec.savings_pct / 100,
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row_idx, col_idx, val)
            if col_idx in (3, 4, 5):
                cell.number_format = _CURRENCY_FMT
                _style(cell, font=_font(10), fill=fill,
                       align=_align("right", "center"))
            elif col_idx == 6:
                cell.number_format = "0.0%"
                _style(cell, font=_font(10), fill=fill,
                       align=_align("right", "center"))
            else:
                _style(cell, font=_font(10), fill=fill,
                       align=_align("left", "center"))

    # Totals row
    total_row = len(recommendations) + 2
    ws.row_dimensions[total_row].height = 17
    total_payg = sum(r.payg_compute_monthly for r in recommendations)
    total_rsv  = sum(r.rsv_compute_monthly  for r in recommendations)
    total_save = total_payg - total_rsv
    total_pct  = (total_save / total_payg) if total_payg else 0

    totals = ["TOTAL", len(recommendations), total_payg, total_rsv, total_save, total_pct]
    for col_idx, val in enumerate(totals, 1):
        cell = ws.cell(total_row, col_idx, val)
        if col_idx in (3, 4, 5):
            cell.number_format = _CURRENCY_FMT
            _style(cell, font=_font(11, bold=True), fill=_fill(_GRAY_TOTAL),
                   align=_align("right", "center"))
        elif col_idx == 6:
            cell.number_format = "0.0%"
            _style(cell, font=_font(11, bold=True), fill=_fill(_GRAY_TOTAL),
                   align=_align("right", "center"))
        else:
            _style(cell, font=_font(11, bold=True), fill=_fill(_GRAY_TOTAL),
                   align=_align("left", "center"))

    # ISF note
    note_row = total_row + 2
    ws.row_dimensions[note_row].height = 30
    ws.cell(note_row, 1,
            "Azure Instance Size Flexibility (ISF) may extend reservation coverage "
            "across sizes within the same VM family. "
            f"Disk costs (${total_disk_monthly:,.2f}/mo) are identical for PAYG and Reserved.")
    _style(ws.cell(note_row, 1),
           font=_font(10, italic=True),
           fill=_fill(_GRAY_DISC),
           align=_align("left", "center", wrap=True))
    ws.merge_cells(f"A{note_row}:{get_column_letter(len(headers))}{note_row}")


# ---------------------------------------------------------------------------
# Sheet 4 — Summary
# ---------------------------------------------------------------------------

def _build_summary_sheet(
    wb: Workbook,
    results: list[VMResult],
    pricing_mode: str,
    region: str,
    disk_type: str,
    currency: str,
    reserved_term: str = "3-year",
    hybrid_benefit: bool = False,
    realistic_top_skus: frozenset[str] = frozenset(),
) -> None:
    ws = wb.create_sheet("Summary")
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 22

    region_display = _region_display(region)
    powered_on  = [r for r in results if r.vm.is_powered_on]
    powered_off = [r for r in results if not r.vm.is_powered_on]

    total_vcpus     = sum(r.vm.vcpus for r in powered_on)
    total_ram_gb    = sum(r.vm.ram_gb for r in powered_on)
    total_disks     = sum(len(r.vm.effective_disks) for r in powered_on)
    total_disk_mo   = sum(r.disk_monthly for r in powered_on)
    total_payg_comp = sum(r.payg_compute_monthly for r in powered_on)
    total_payg      = total_payg_comp + total_disk_mo

    if pricing_mode == "realistic":
        total_rsv_comp = sum(
            (r.reserved_compute_monthly
             if r.sku and r.sku.name in realistic_top_skus and r.reserved_compute_monthly is not None
             else r.payg_compute_monthly)
            for r in powered_on if r.sku is not None
        )
    else:
        rsv_eligible = [r for r in powered_on
                        if r.sku and r.reserved_compute_monthly is not None]
        total_rsv_comp = sum(r.reserved_compute_monthly for r in rsv_eligible)  # type: ignore[misc]
    total_rsv = total_rsv_comp + total_disk_mo

    show_reserved = pricing_mode in ("all-reserved", "realistic")

    from .output import _rsv_label_long
    rsv_long = _rsv_label_long(reserved_term)
    yrs = "3-yr" if reserved_term == "3-year" else "1-yr"

    if pricing_mode == "realistic":
        mode_label = f"Realistic (top {len(realistic_top_skus)} SKUs at {rsv_long}, rest PAYG)"
    elif pricing_mode == "all-reserved":
        mode_label = f"All {rsv_long}"
    else:
        mode_label = "All Pay-as-you-go"
    if hybrid_benefit:
        mode_label += " + Azure Hybrid Benefit"

    rsv_compute_label = f"{rsv_long} compute/mo"
    rsv_total_label   = f"Total {rsv_long}/mo"

    rows = [
        ("Report generated",  datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")),
        ("Region",            region_display),
        ("Pricing mode",      mode_label),
        ("Disk type",         disk_type),
        (None, None),
        ("VMs (powered-on)",  len(powered_on)),
        ("VMs (powered-off)", len(powered_off)),
        ("Total vCPUs",       total_vcpus),
        ("Total RAM (GB)",    round(total_ram_gb, 1)),
        ("Total disks",       total_disks),
        (None, None),
        ("PAYG compute/mo",   total_payg_comp),
        ("Disk cost/mo",      total_disk_mo),
        ("Total PAYG/mo",     total_payg),
    ]
    if show_reserved:
        rows += [
            (None, None),
            (rsv_compute_label,           total_rsv_comp),
            ("Disk cost/mo (unchanged)",  total_disk_mo),
            (rsv_total_label,             total_rsv),
            ("Compute savings/mo",        total_payg_comp - total_rsv_comp),
            ("Savings %",                 (total_payg_comp - total_rsv_comp) / total_payg_comp
             if total_payg_comp else 0),
        ]

    money_labels = {
        "PAYG compute/mo", "Disk cost/mo", "Total PAYG/mo",
        rsv_compute_label, "Disk cost/mo (unchanged)",
        rsv_total_label, "Compute savings/mo",
    }
    pct_labels = {"Savings %"}

    # Title
    ws.row_dimensions[1].height = 22
    ws.cell(1, 1, "Azure IaaS Estimate — Summary")
    _style(ws.cell(1, 1),
           font=_font(13, bold=True, color=_BLUE_TITLE),
           align=_align("left", "center"))
    ws.merge_cells("A1:B1")

    for row_idx, (label, value) in enumerate(rows, 2):
        ws.row_dimensions[row_idx].height = 15
        if label is None:
            continue
        label_cell = ws.cell(row_idx, 1, label)
        value_cell = ws.cell(row_idx, 2, value)

        _style(label_cell, font=_font(11, bold=True), align=_align("left", "center"))

        if label in money_labels and isinstance(value, (int, float)):
            value_cell.number_format = _CURRENCY_FMT
            _style(value_cell, font=_font(11), align=_align("right", "center"))
        elif label in pct_labels and isinstance(value, (int, float)):
            value_cell.number_format = "0.0%"
            _style(value_cell, font=_font(11), align=_align("right", "center"))
        elif isinstance(value, (int, float)) and label not in ("VMs (powered-on)",
                                                                "VMs (powered-off)",
                                                                "Total vCPUs",
                                                                "Total disks"):
            value_cell.number_format = _INT_FMT
            _style(value_cell, font=_font(11), align=_align("right", "center"))
        else:
            _style(value_cell, font=_font(11), align=_align("left", "center"))


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def write_workbook(
    results: list[VMResult],
    recommendations: list[ReservationRec],
    output_path: str | Path,
    region: str,
    currency: str,
    pricing_mode: str,
    disk_type: str,
    disk_source: str = "provisioned",
    rvtools_filename: str = "",
    reserved_term: str = "3-year",
    hybrid_benefit: bool = False,
    support_plan: str = "basic",
    realistic_top_skus: frozenset[str] = frozenset(),
) -> None:
    """
    Write an Azure Calculator-style Excel workbook to *output_path*.

    Sheets:
      1. Estimate   — Azure Calculator look-alike summary
      2. VM Detail  — per-VM pricing table
      3. Reservations — reservation recommendations (always included)
      4. Summary    — aggregate statistics
    """
    wb = Workbook()

    _build_estimate_sheet(
        wb, results, recommendations, region, currency,
        pricing_mode, disk_type, rvtools_filename,
        reserved_term=reserved_term, hybrid_benefit=hybrid_benefit,
        support_plan=support_plan, realistic_top_skus=realistic_top_skus,
        disk_source=disk_source,
    )
    _build_vm_detail_sheet(wb, results, pricing_mode, region, reserved_term=reserved_term)

    total_disk = sum(r.disk_monthly for r in results if r.vm.is_powered_on)
    _build_reservations_sheet(wb, recommendations, total_disk, reserved_term=reserved_term)

    _build_summary_sheet(wb, results, pricing_mode, region, disk_type, currency,
                         reserved_term=reserved_term, hybrid_benefit=hybrid_benefit,
                         realistic_top_skus=realistic_top_skus)

    wb.save(output_path)
